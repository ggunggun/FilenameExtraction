# __author__ = 'gengt'

from openpyxl.workbook import Workbook
#from openpyxl.writer.excel import ExcelWriter as EWriter


class ExcelWriter(object):

    def __init__(self, fileName, workBook=Workbook()):
        self.fileName = fileName
        self.workBook = workBook
        #self.excelWriter = EWriter(workbook = self.workBook)

    def save(self):
        #excelWriter = EWriter(workbook=workBook)
        self.workBook.save(filename=self.fileName)

    def buildWorkBook(self, tables):
        if tables:
            self.workBook = Workbook()
            for sheetName in self.workBook.sheetnames:
                self.workBook.remove(self.workBook[sheetName])
            for name, table in tables.items():
                workSheet = self.workBook.create_sheet(title=name)
                for row in table:
                    workSheet.append(row)